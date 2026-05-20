"""
Email Verifier API — FastAPI backend for SMTP email verification.
Accepts a list of emails (JSON or file upload) and returns verification results.
"""

import asyncio
import csv
import io
import json
import re
import socket
import smtplib
from concurrent.futures import ThreadPoolExecutor
from typing import Optional

import dns.resolver
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse

app = FastAPI(title="Email Verifier")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SENDER = "verify@mailcheck.app"
TIMEOUT = 10
MAX_EMAILS = 500
MAX_WORKERS = 10

executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)


def get_mx(domain: str) -> list[str]:
    """Return MX hosts sorted by priority."""
    try:
        answers = dns.resolver.resolve(domain, "MX")
        mx_hosts = sorted(answers, key=lambda r: r.preference)
        return [str(r.exchange).rstrip(".") for r in mx_hosts]
    except Exception:
        return []


def verify_email(email: str) -> dict:
    """Check if an email is accepted by the recipient's SMTP server."""
    email = email.strip().lower()
    if not re.match(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$", email):
        return {"email": email, "status": "invalid", "detail": "Invalid email format"}

    domain = email.split("@")[1]
    result = {"email": email, "status": "unknown", "detail": ""}

    mx_hosts = get_mx(domain)
    if not mx_hosts:
        result["status"] = "error"
        result["detail"] = f"No MX records found for {domain}"
        return result

    for mx in mx_hosts:
        try:
            smtp = smtplib.SMTP(timeout=TIMEOUT)
            smtp.connect(mx, 25)
            smtp.helo(socket.getfqdn())
            smtp.mail(SENDER)
            code, msg = smtp.rcpt(email)
            smtp.quit()

            msg_str = msg.decode(errors="ignore")
            if code == 250:
                result["status"] = "valid"
                result["detail"] = f"Accepted by {mx}"
            elif code in (550, 553):
                result["status"] = "invalid"
                result["detail"] = f"Rejected by {mx}: {msg_str}"
            else:
                result["status"] = "uncertain"
                result["detail"] = f"{mx} returned {code}: {msg_str}"
            return result

        except smtplib.SMTPServerDisconnected:
            result["detail"] = f"{mx}: server disconnected"
        except smtplib.SMTPConnectError as e:
            result["detail"] = f"{mx}: connection error"
        except socket.timeout:
            result["detail"] = f"{mx}: timeout"
        except Exception as e:
            result["detail"] = f"{mx}: {e}"

    result["status"] = "error"
    result["detail"] = f"Could not connect to any MX for {domain}"
    return result


def parse_emails_from_text(text: str) -> list[str]:
    """Extract emails from text — supports comma, newline, semicolon separated."""
    emails = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    return list(dict.fromkeys(emails))  # deduplicate preserving order


def parse_emails_from_csv(content: str) -> list[str]:
    """Extract emails from CSV content."""
    emails = []
    reader = csv.reader(io.StringIO(content))
    for row in reader:
        for cell in row:
            found = re.findall(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", cell)
            emails.extend(found)
    return list(dict.fromkeys(emails))


@app.post("/api/verify")
async def verify_emails_endpoint(
    emails: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None),
):
    """Verify a list of emails. Accepts either a text list or file upload."""
    email_list = []

    if file:
        content = await file.read()
        filename = file.filename.lower()

        if filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            found = re.findall(
                                r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}",
                                str(cell),
                            )
                            email_list.extend(found)
            wb.close()
        elif filename.endswith(".csv"):
            text = content.decode("utf-8", errors="ignore")
            email_list = parse_emails_from_csv(text)
        else:
            text = content.decode("utf-8", errors="ignore")
            email_list = parse_emails_from_text(text)

    elif emails:
        email_list = parse_emails_from_text(emails)
    else:
        raise HTTPException(status_code=400, detail="No emails provided")

    email_list = list(dict.fromkeys(email_list))  # deduplicate

    if len(email_list) > MAX_EMAILS:
        raise HTTPException(
            status_code=400,
            detail=f"Too many emails. Maximum is {MAX_EMAILS}, got {len(email_list)}",
        )

    if not email_list:
        raise HTTPException(status_code=400, detail="No valid emails found in input")

    # Run verifications in thread pool for concurrency
    loop = asyncio.get_event_loop()
    results = await asyncio.gather(
        *[loop.run_in_executor(executor, verify_email, e) for e in email_list]
    )

    summary = {
        "total": len(results),
        "valid": sum(1 for r in results if r["status"] == "valid"),
        "invalid": sum(1 for r in results if r["status"] == "invalid"),
        "uncertain": sum(1 for r in results if r["status"] == "uncertain"),
        "error": sum(1 for r in results if r["status"] == "error"),
    }

    return {"results": results, "summary": summary}


# Serve static files
app.mount("/", StaticFiles(directory="static", html=True), name="static")
